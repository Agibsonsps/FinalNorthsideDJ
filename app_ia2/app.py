import openpyxl
from flask import Flask, render_template, request, redirect, url_for, flash, session
import sqlite3
import requests
import pandas as pd
from werkzeug.security import generate_password_hash, check_password_hash
app = Flask(__name__)
app.secret_key = 's3cr3t'


def hash_password(password):
    # Hash the password using bcrypt for secure storage
    return generate_password_hash(password, method='pbkdf2:sha256')


gamesdoc = pd.read_excel('/Users/asgibsonpc2022/PycharmProjects/FinalNorthsideDJ/app_ia2/GameData.xlsx')
# this is the data for the games


@app.route('/toggle_participation/<int:tournament_id>', methods=['POST'])
def toggle_participation(tournament_id):
    # Check if the user is logged in
    if "user" not in session:
        return redirect(url_for('login'))
    # Get the user ID from the session
    user_id = session["user"]
    # Connect to the database
    with sqlite3.connect('Esportsapp.db') as db:
        cursor = db.cursor()
        cursor.execute(f'SELECT 1 FROM tournament_players_{tournament_id} WHERE user_ID = ?', (user_id,))
        participation = cursor.fetchone()
# Check if the user is already participating in the tournament
        if participation:
            cursor.execute(f'DELETE FROM tournament_players_{tournament_id} WHERE user_ID = ?', (user_id,))
        else:
            cursor.execute(f'INSERT INTO tournament_players_{tournament_id} (tournament_ID, user_ID) VALUES (?, ?)', (tournament_id, user_id))
# Commit the changes to the database
        db.commit()
# Redirect to the previous page
    return redirect(request.referrer or url_for('index'))


@app.route('/')
def index():
    if "user" not in session:
        return redirect(url_for('login'))
    user_id = session['user']
    with sqlite3.connect('Esportsapp.db') as db:
        cursor = db.cursor()
        cursor.execute('SELECT username FROM users WHERE user_ID = ?', (user_id,))
        username = cursor.fetchone()[0]
# Fetch all tournaments from the database
        cursor.execute('SELECT * FROM tournament')
        tournaments = cursor.fetchall()
        tournaments_with_participation = []
# Fetch the players for each tournament
        for tournament in tournaments:
            tournament_id = tournament[0]
# Fetch the players for the tournament
            cursor.execute(f'''
                SELECT u.firstname 
                FROM tournament_players_{tournament_id} tp
                JOIN users u ON tp.user_ID = u.user_ID
            ''')
            players = [row[0] for row in cursor.fetchall()]
# Check if the user is participating in the tournament
            cursor.execute(f'SELECT 1 FROM tournament_players_{tournament_id} WHERE user_ID = ?', (user_id,))
            participating = cursor.fetchone() is not None
# Append the tournament data to the list
            tournaments_with_participation.append((tournament, players, participating))
# Render the index template with the tournament data
    return render_template('index.html', username=username,
                           tournaments_with_participation=tournaments_with_participation)


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        # Get the user data from the form
        firstname = request.form['firstname']
        lastname = request.form['firstname']
        user_type = request.form['user_type']
        email = request.form['email']
        username = request.form['username']
        password = request.form['password']
        # Hash the password using the hash_password function
        hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
        # Connect to the database
        try:
            # Insert the user data into the database
            with sqlite3.connect('Esportsapp.db') as db:
                # Create a new cursor
                cursor = db.cursor()
                # Insert user data into the 'users' table, using parameters to prevent SQL injection
                cursor.execute('INSERT INTO users (firstname, lastname, user_type, email, username, hashed_password) VALUES (?, ?, ?, ?, ?, ?)', (firstname, lastname, user_type, email, username, hashed_password))
                db.commit()
            flash('Registered successfully!', 'success')
            # Redirect to the login page
            return redirect(url_for('login'))
        except sqlite3.IntegrityError:
            flash('Username already exists!', 'danger')
            # Redirect to the register page
    return render_template('register.html')


@app.route('/login', methods=['GET', 'POST'])
def login():

    if request.method == 'POST':
        # Get the username and password from the form
        username = request.form['username']
        password = request.form['password']
        with sqlite3.connect('Esportsapp.db') as db:
            cursor = db.cursor()
            # Select user data using a parameter to prevent SQL injection
            # Fetch the user data from the database
            cursor.execute('SELECT * FROM users WHERE username = ?', (username,))
            user = cursor.fetchone()
            cursor.close()
            # Check if the user exists and the password is correct
            if user and check_password_hash(user[6], password):
                session['user'] = user[0]
                # Redirect to the index page
                return redirect(url_for('index'))
            else:
                flash('Invalid username or password!', 'danger')
    return render_template('login.html')


@app.route('/logout')
def logout():
    # Remove the user from the session
    session.pop('user', None)
    return redirect(url_for('index'))


@app.route('/search', methods=['GET', 'POST'])
def search():
    # Check if the user is logged in
    if "user" not in session:
        return redirect(url_for('login'))
    # Check if the form was submitted
    if request.method == 'POST':
        # Get the query from the form data
        query = request.form.get('query')
        gameresults = search_games(query)
        # Render the results template with the query and game results
        return render_template('results.html', query=query, gameresults=gameresults)
    return render_template('search.html')


@app.route('/results', methods=['GET', 'POST'])
def results():
    if "user" not in session:
        return redirect(url_for('login'))
# Get the user ID from the session
    user_id = session['user']
    query = request.args.get('query')
    gameresults = search_games(query)

    # Fetch the user's favorite games
    with sqlite3.connect('Esportsapp.db') as db:
        cursor = db.cursor()
        cursor.execute('SELECT game_ID FROM fave_games WHERE user_ID = ?', (user_id,))
        favorite_games = [row[0] for row in cursor.fetchall()]
# Render the results template with the query, game results, and favorite games
    return render_template('results.html', query=query, gameresults=gameresults, favorite_games=favorite_games)


@app.route('/toggle_favorite', methods=['POST'])
def toggle_favorite():
    if "user" not in session:
        return redirect(url_for('login'))

    user_id = session['user']
    game_id = request.form.get('game_id')
    action = request.form.get('action')
    query = request.form.get('query') # Fetching the game_query from the form data
    with sqlite3.connect('Esportsapp.db') as db:
        cursor = db.cursor()
        # Check if the user has already favorited the game
        if action == 'favorite':
            cursor.execute('INSERT INTO fave_games (user_ID, game_ID) VALUES (?, ?)', (user_id, game_id))
        # If the action is 'unfavorite', delete the favorite game from the database
        elif action == 'unfavorite':
            cursor.execute('DELETE FROM fave_games WHERE user_ID = ? AND game_ID = ?', (user_id, game_id))
        db.commit()

    # Redirect to the results route with the game_query parameter
    return redirect(url_for('results', query=query))


def search_games(query):
    # Load the game data from the Excel file
    query = query.lower()
    wb = openpyxl.load_workbook('/Users/asgibsonpc2022/PycharmProjects/FinalNorthsideDJ/app_ia2/GameData.xlsx')
    ws = wb["Data_Used_For_Reviews"]
    # Create a dictionary to store the game data
    games_dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Check if the YouTube trailer link is empty
        if not row[25]:
            # Skip the row if the trailer link is empty
            continue
            # Get the game title from the row
        game_title = row[2]
        # Create a dictionary with the game details
        game_details = {
            "trailer_link": row[25],
            "thumbnail_link": f"https://img.youtube.com/vi/{row[25].split('v=')[1]}/0.jpg" if "youtube.com" in row[25] else "",
            "id": row[1],
        }
        # Check if the query is in the game title
        if query in game_title.lower():
            games_dict[game_title] = game_details
    return games_dict


@app.route('/profile', methods=['GET', 'POST'])
def profile():
    if "user" not in session:
        return redirect(url_for('login'))
    # Get the user ID from the session
    user_id = session['user']
    # Fetch the user data from the database
    with sqlite3.connect('Esportsapp.db') as db:
        cursor = db.cursor()
        # Select the user data using a parameter to prevent SQL injection
        cursor.execute('SELECT username FROM users WHERE user_ID = ?', (user_id,))
        username = cursor.fetchone()[0]
        cursor.execute('SELECT email FROM users WHERE user_ID = ?', (user_id,))
        email = cursor.fetchone()[0]
        cursor.execute('SELECT user_type FROM users WHERE user_ID = ?', (user_id,))
        user_type = cursor.fetchone()[0]
        cursor.execute('SELECT firstname FROM users WHERE user_ID = ?', (user_id,))
        firstname = cursor.fetchone()[0]
        cursor.execute('SELECT lastname FROM users WHERE user_ID = ?', (user_id,))
        lastname = cursor.fetchone()[0]
        cursor.close()
        # Fetch the user's favorite games
    favorite_games_ids = []
    favourite_games_data = []

    with sqlite3.connect('Esportsapp.db') as db:
        cursor = db.cursor()
        # Select the favorite games using a parameter to prevent SQL injection
        cursor.execute('SELECT game_ID FROM fave_games WHERE user_ID = ?', (user_id,))
        favorite_games_ids = [row[0] for row in cursor.fetchall()]
        cursor.close()
# Load the game data from the Excel file
    wb = openpyxl.load_workbook('/Users/asgibsonpc2022/PycharmProjects/FinalNorthsideDJ/app_ia2/GameData.xlsx')
    ws = wb["Data_Used_For_Reviews"]
# Iterate over the rows in the Excel sheet
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] in favorite_games_ids:
            # Append the game data to the list
            favourite_games_data.append({
                "id": row[1],
                "title": row[2],
                "trailer_link": row[25]
            })
    return render_template('profile.html', favorite_games_ids=favorite_games_ids, favourite_games_data=favourite_games_data, username=username, email=email, user_type=user_type, first_name=firstname, last_name=lastname, user_id=user_id)


@app.route('/add_favorite', methods=['POST'])
def add_favorite():
    if "user" not in session:
        return redirect(url_for('login'))
    # Get the game ID and title from the form data
    game_id = request.form['game_id']
    game_title = request.form['game_title']
    user_id = session['user']
    with sqlite3.connect('Esportsapp.db') as db:
        cursor = db.cursor()
        # Select the favorite games for the user
        cursor.execute('SELECT * FROM fave_games WHERE user_ID = ?', (user_id,))
        favorite_games = cursor.fetchall()
        db.cursor().close()
    if game_id in favorite_games:
        flash('Game already in favorites!', 'danger')
        return redirect(url_for('results'))
    else:
        # Insert the favorite game into the database
        with sqlite3.connect('Esportsapp.db') as db:
            cursor = db.cursor()
            cursor.execute('INSERT INTO fave_games (user_ID, game_ID) VALUES (?, ?)', (user_id, game_id))
            db.commit()
            # Flash a success message
            flash(game_title + ' added to favorites!')
            return redirect(url_for('results'))


@app.route('/tournament', methods=['GET', 'POST'])
def tournament():
    if "user" not in session:
        return redirect(url_for('login'))
# Get the user ID from the session
    user_id = session["user"]
    tournaments_with_participation = []
    # Fetch all tournaments from the database
    with sqlite3.connect('Esportsapp.db') as db:
        cursor = db.cursor()
        cursor.execute('SELECT * FROM tournament')
        tournaments = cursor.fetchall()
# Fetch the players for each tournament
        if not tournaments:
            flash('No tournaments available!', 'info')
            return render_template('tournament.html')
# Fetch the players for each tournament
        for tournament in tournaments:
            tournament_id = tournament[0]
# Fetch the players for the tournament
            cursor.execute(f'''
                SELECT u.firstname 
                FROM tournament_players_{tournament_id} tp
                JOIN users u ON tp.user_ID = u.user_ID
            ''')
            # Fetch the players for the tournament
            players = [row[0] for row in cursor.fetchall()]
# Check if the user is participating in the tournament
            cursor.execute(f'SELECT 1 FROM tournament_players_{tournament_id} WHERE user_ID = ?', (user_id,))
            participating = cursor.fetchone() is not None

            tournaments_with_participation.append((tournament, players, participating))

    return render_template('tournament.html', tournaments=tournaments_with_participation)


@app.route('/create_tournament', methods=['GET', 'POST'])
def create_tournament():
    if "user" not in session:
        return redirect(url_for('login'))
# Get the user ID from the session
    game_tn_url = ''
    if request.method == 'POST':
        # Get the tournament data from the form
        type = request.form.get('type')
        start = request.form.get('start')
        status = request.form.get('status')
        description = request.form.get('description')
        end = request.form.get('end')
        game_ID = request.form.get('game_ID')
        platform = request.form.get('platform')
# Connect to the database
        wb = openpyxl.load_workbook('/Users/asgibsonpc2022/PycharmProjects/FinalNorthsideDJ/app_ia2/GameData.xlsx')
        ws = wb["Data_Used_For_Reviews"]
# Iterate over the rows in the Excel sheet
        try:
            game_ID = int(game_ID)
        except ValueError:
            flash('Invalid game ID!', 'danger')
            return redirect(url_for('create_tournament'))
# Check if the game ID is valid
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row[1] == game_ID:
                game_tn_url = row[25]
                break
# Insert the tournament data into the database
        with sqlite3.connect('Esportsapp.db') as db:
            cursor = db.cursor()
            cursor.execute('''
                INSERT INTO tournament (type, start, status, description, end, game_ID, platform, game_tn_url)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (type, start, status, description, end, game_ID, platform, game_tn_url))
# Get the ID of the inserted tournament
            cursor.execute('SELECT last_insert_rowid()')
            tournament_id = cursor.fetchone()[0]
# Create a new table for the tournament players
            cursor.execute(f'''
                CREATE TABLE IF NOT EXISTS tournament_players_{tournament_id} (
                    player_ID INTEGER PRIMARY KEY AUTOINCREMENT,
                    tournament_ID INTEGER,
                    user_ID INTEGER
                )
            ''')
# Commit the changes to the database
            db.commit()

        return redirect(url_for('tournament'))

    return render_template('create_tournament.html')


@app.route('/event_data')
def event_data():
    if "user" not in session:
        return redirect(url_for('login'))
    with sqlite3.connect('Esportsapp.db') as db:
        cursor = db.cursor()
        cursor.execute('''SELECT * FROM events''')
        # Fetch all events from the database
        events = cursor.fetchall()
        # Render the event_data template with the event data
    return render_template('event_data.html', events=events)


if __name__ == '__main__':
    app.run(debug=True, port=4002)


